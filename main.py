# Sistema de Controle de Estoque
# Exercício 04 - Programação II
# Trabalho feito para gerenciar estoque usando Excel

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os


def criar_planilha_estoque(nome_arquivo):
    # cria a planilha com as 3 abas necessarias
    try:
        wb = Workbook()

        # aba de produtos
        ws_produtos = wb.active
        ws_produtos.title = "Produtos"
        criar_cabecalho_produtos(ws_produtos)

        # aba de movimentacoes
        ws_movimentacoes = wb.create_sheet("Movimentações")
        criar_cabecalho_movimentacoes(ws_movimentacoes)

        # aba de relatorios
        ws_relatorios = wb.create_sheet("Relatórios")
        criar_cabecalho_relatorios(ws_relatorios)

        wb.save(nome_arquivo)
        print(f"Planilha '{nome_arquivo}' criada com sucesso!")
        return True

    except PermissionError:
        print(f"ERRO: Não foi possível salvar o arquivo. Verifique se ele está aberto.")
        return False


def criar_cabecalho_produtos(ws):
    # cabecalho da aba produtos
    cabecalhos = ["Código", "Nome do Produto", "Categoria", "Quantidade",
                  "Estoque Mínimo", "Preço Unitário", "Valor Total", "Status"]

    for col, titulo in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=1, column=col)
        celula.value = titulo
        celula.font = Font(bold=True)
        celula.alignment = Alignment(horizontal="center")

    # ajustar tamanho das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12


def criar_cabecalho_movimentacoes(ws):
    # cabecalho da aba movimentacoes
    cabecalhos = ["Data", "Código Produto", "Tipo", "Quantidade", "Responsável", "Observações"]

    for col, titulo in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=1, column=col)
        celula.value = titulo
        celula.font = Font(bold=True)
        celula.alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30


def criar_cabecalho_relatorios(ws):
    # configurar aba de relatorios
    ws['A1'] = "RELATÓRIO DE ESTOQUE"
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = "Estatísticas Gerais:"
    ws['A3'].font = Font(bold=True)

    # informacoes que vao aparecer
    ws['A4'] = "Total de Produtos Cadastrados:"
    ws['A5'] = "Valor Total do Estoque:"
    ws['A6'] = "Produtos com Estoque Baixo:"
    ws['A7'] = "Data do Relatório:"

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def adicionar_produto(nome_arquivo, codigo, nome, categoria, quantidade, estoque_minimo, preco):
    # adiciona produto novo na planilha
    try:
        wb = load_workbook(nome_arquivo)
        ws = wb["Produtos"]

        # verificar se ja existe produto com esse codigo
        produtos_existentes = []
        for row in range(2, ws.max_row + 1):
            cod = ws.cell(row=row, column=1).value
            if cod:
                produtos_existentes.append(cod)

        if codigo in produtos_existentes:
            print(f"Erro: Produto com código '{codigo}' já existe!")
            return False

        # achar proxima linha vazia
        proxima_linha = ws.max_row + 1

        # colocar os dados do produto
        ws.cell(row=proxima_linha, column=1).value = codigo
        ws.cell(row=proxima_linha, column=2).value = nome
        ws.cell(row=proxima_linha, column=3).value = categoria
        ws.cell(row=proxima_linha, column=4).value = quantidade
        ws.cell(row=proxima_linha, column=5).value = estoque_minimo
        ws.cell(row=proxima_linha, column=6).value = preco

        # formula para calcular valor total
        ws.cell(row=proxima_linha, column=7).value = f"=D{proxima_linha}*F{proxima_linha}"

        # formula para ver se o estoque ta baixo
        ws.cell(row=proxima_linha, column=8).value = \
            f'=SE(D{proxima_linha}<E{proxima_linha},"BAIXO","OK")'

        # formatar preco como R$
        ws.cell(row=proxima_linha, column=6).number_format = 'R$ #,##0.00'
        ws.cell(row=proxima_linha, column=7).number_format = 'R$ #,##0.00'

        wb.save(nome_arquivo)
        print(f"Produto '{nome}' adicionado com sucesso!")
        return True

    except FileNotFoundError:
        print(f"ERRO: Arquivo '{nome_arquivo}' não encontrado!")
        return False
    except PermissionError:
        print(f"ERRO: Arquivo está aberto. Feche o Excel e tente novamente.")
        return False


def registrar_movimentacao(nome_arquivo, codigo_produto, tipo, quantidade):
    # registra entrada ou saida de produtos no estoque
    try:
        wb = load_workbook(nome_arquivo)
        ws_produtos = wb["Produtos"]
        ws_movimentacoes = wb["Movimentações"]

        # procurar o produto pelo codigo
        linha_produto = None
        for row in range(2, ws_produtos.max_row + 1):
            if ws_produtos.cell(row=row, column=1).value == codigo_produto:
                linha_produto = row
                break

        if linha_produto is None:
            print(f"ERRO: Produto com código '{codigo_produto}' não encontrado!")
            return False

        # pegar quantidade atual do produto
        quantidade_atual = ws_produtos.cell(row=linha_produto, column=4).value
        if quantidade_atual is None:
            quantidade_atual = 0

        if tipo.upper() == "ENTRADA":
            nova_quantidade = quantidade_atual + quantidade
        elif tipo.upper() == "SAÍDA":
            # verificar se tem produto suficiente
            if quantidade_atual < quantidade:
                print(f"Erro: Estoque insuficiente! Disponível: {quantidade_atual}")
                return False
            nova_quantidade = quantidade_atual - quantidade
        else:
            print(f"Erro: Tipo de movimentação inválido! Use 'ENTRADA' ou 'SAÍDA'")
            return False

        # atualizar a quantidade do produto
        ws_produtos.cell(row=linha_produto, column=4).value = nova_quantidade

        # registrar a movimentacao na outra aba
        proxima_linha = ws_movimentacoes.max_row + 1
        ws_movimentacoes.cell(row=proxima_linha, column=1).value = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws_movimentacoes.cell(row=proxima_linha, column=2).value = codigo_produto
        ws_movimentacoes.cell(row=proxima_linha, column=3).value = tipo.upper()
        ws_movimentacoes.cell(row=proxima_linha, column=4).value = quantidade
        ws_movimentacoes.cell(row=proxima_linha, column=5).value = "Sistema"
        ws_movimentacoes.cell(row=proxima_linha, column=6).value = "-"

        wb.save(nome_arquivo)
        print(f"Movimentação de {tipo} registrada com sucesso!")
        print(f"Quantidade anterior: {quantidade_atual} | Nova quantidade: {nova_quantidade}")
        return True

    except FileNotFoundError:
        print(f"ERRO: Arquivo '{nome_arquivo}' não encontrado!")
        return False
    except PermissionError:
        print(f"ERRO: Arquivo está aberto. Feche o Excel e tente novamente.")
        return False


def atualizar_relatorio(nome_arquivo):
    # atualiza a aba de relatorios com as estatisticas

    # verificar se o arquivo existe
    if not os.path.exists(nome_arquivo):
        print(f"ERRO: Arquivo '{nome_arquivo}' não encontrado!")
        print("Adicione produtos primeiro para criar a planilha.")
        return False

    try:
        wb = load_workbook(nome_arquivo)
        ws_relatorios = wb["Relatórios"]
        ws_produtos = wb["Produtos"]

        # contar quantos produtos tem
        num_produtos = ws_produtos.max_row - 1
        if num_produtos < 0:
            num_produtos = 0

        # colocar o numero de produtos
        ws_relatorios['B4'] = num_produtos

        # formula para calcular valor total do estoque
        if num_produtos > 0:
            ws_relatorios['B5'] = f"=SOMA(Produtos!G2:G{ws_produtos.max_row})"
            ws_relatorios['B5'].number_format = 'R$ #,##0.00'
        else:
            ws_relatorios['B5'] = "R$ 0,00"

        # contar produtos com estoque baixo
        if num_produtos > 0:
            ws_relatorios['B6'] = f'=CONT.SE(Produtos!H2:H{ws_produtos.max_row},"BAIXO")'
        else:
            ws_relatorios['B6'] = 0

        # data e hora do relatorio
        ws_relatorios['B7'] = datetime.now().strftime("%d/%m/%Y %H:%M")

        # fazer tabela de produtos com estoque baixo
        ws_relatorios['A9'] = "Produtos com Estoque Baixo:"
        ws_relatorios['A9'].font = Font(bold=True)

        ws_relatorios['A10'] = "Código"
        ws_relatorios['B10'] = "Nome"
        ws_relatorios['C10'] = "Quantidade Atual"
        ws_relatorios['D10'] = "Estoque Mínimo"

        # formatar cabecalho da tabela
        for col in range(1, 5):
            celula = ws_relatorios.cell(row=10, column=col)
            celula.font = Font(bold=True)

        # procurar produtos com estoque baixo
        linha_relatorio = 11
        for row in range(2, ws_produtos.max_row + 1):
            codigo = ws_produtos.cell(row=row, column=1).value
            if codigo is None:
                continue

            quantidade = ws_produtos.cell(row=row, column=4).value
            estoque_minimo = ws_produtos.cell(row=row, column=5).value

            if quantidade is None:
                quantidade = 0
            if estoque_minimo is None:
                estoque_minimo = 0

            # se tiver abaixo do minimo, adiciona na tabela
            if quantidade < estoque_minimo:
                ws_relatorios.cell(row=linha_relatorio, column=1).value = codigo
                ws_relatorios.cell(row=linha_relatorio, column=2).value = \
                    ws_produtos.cell(row=row, column=2).value
                ws_relatorios.cell(row=linha_relatorio, column=3).value = quantidade
                ws_relatorios.cell(row=linha_relatorio, column=4).value = estoque_minimo


                linha_relatorio += 1

        wb.save(nome_arquivo)
        print(f"Relatório atualizado com sucesso!")
        return True

    except FileNotFoundError:
        print(f"ERRO: Arquivo '{nome_arquivo}' não encontrado!")
        return False
    except PermissionError:
        print(f"ERRO: Arquivo está aberto. Feche o Excel e tente novamente.")
        return False


def listar_produtos(nome_arquivo):
    # mostra todos os produtos que tem no estoque

    # verificar se o arquivo existe
    if not os.path.exists(nome_arquivo):
        print(f"\nERRO: Arquivo '{nome_arquivo}' não encontrado!")
        print("Adicione produtos primeiro para criar a planilha.")
        return

    try:
        wb = load_workbook(nome_arquivo)
        ws = wb["Produtos"]

        print("\n" + "="*80)
        print("LISTA DE PRODUTOS EM ESTOQUE")
        print("="*80)

        # ver se tem algum produto
        if ws.max_row < 2:
            print("Nenhum produto cadastrado.")
            return

        # mostrar cabecalho da tabela
        print(f"{'Código':<10} {'Nome':<25} {'Categoria':<15} {'Qtd':<8} {'Preço':<12}")
        print("-"*80)

        # mostrar todos os produtos
        for row in range(2, ws.max_row + 1):
            codigo = ws.cell(row=row, column=1).value
            nome = ws.cell(row=row, column=2).value
            categoria = ws.cell(row=row, column=3).value
            quantidade = ws.cell(row=row, column=4).value
            preco = ws.cell(row=row, column=6).value

            # pular linhas vazias
            if codigo is None:
                continue

            # mostrar o produto
            if categoria is None:
                categoria = "N/A"
            if quantidade is None:
                quantidade = 0
            if preco is None:
                preco = 0.0

            print(f"{codigo:<10} {nome:<25} {categoria:<15} {quantidade:<8} R$ {preco:>8.2f}")

        print("="*80 + "\n")

    except FileNotFoundError:
        print(f"ERRO: Arquivo '{nome_arquivo}' não encontrado!")
    except PermissionError:
        print(f"ERRO: Arquivo está aberto. Feche o Excel e tente novamente.")



def menu_principal():
    # mostra o menu e faz o sistema funcionar
    nome_arquivo = "controle_estoque.xlsx"

    print("\n" + "="*60)
    print("SISTEMA DE CONTROLE DE ESTOQUE")
    print("="*60)

    # ver se o arquivo existe, se nao criar
    if not os.path.exists(nome_arquivo):
        print("\nPrimeiro acesso detectado. Criando planilha...")
        criar_planilha_estoque(nome_arquivo)

    while True:
        print("\n" + "-"*60)
        print("MENU PRINCIPAL")
        print("-"*60)
        print("1. Adicionar novo produto")
        print("2. Registrar entrada de estoque")
        print("3. Registrar saída de estoque")
        print("4. Listar todos os produtos")
        print("5. Atualizar relatórios")
        print("6. Sair")
        print("-"*60)

        try:
            opcao = input("Escolha uma opção: ").strip()

            # adicionar produto
            if opcao == "1":
                print("\n--- ADICIONAR NOVO PRODUTO ---")
                codigo = input("Código do produto: ").strip()
                nome = input("Nome do produto: ").strip()
                categoria = input("Categoria: ").strip()
                quantidade = int(input("Quantidade inicial: "))
                estoque_minimo = int(input("Estoque mínimo: "))
                preco = float(input("Preço unitário (R$): "))

                adicionar_produto(nome_arquivo, codigo, nome, categoria,
                                quantidade, estoque_minimo, preco)

            # registrar entrada
            elif opcao == "2":
                print("\n--- REGISTRAR ENTRADA DE ESTOQUE ---")
                codigo = input("Código do produto: ").strip()
                quantidade = int(input("Quantidade a adicionar: "))

                registrar_movimentacao(nome_arquivo, codigo, "ENTRADA", quantidade)

            # registrar saida
            elif opcao == "3":
                print("\n--- REGISTRAR SAÍDA DE ESTOQUE ---")
                codigo = input("Código do produto: ").strip()
                quantidade = int(input("Quantidade a retirar: "))

                registrar_movimentacao(nome_arquivo, codigo, "SAÍDA", quantidade)

            # listar produtos
            elif opcao == "4":
                listar_produtos(nome_arquivo)

            # atualizar relatorios
            elif opcao == "5":
                print("\nAtualizando relatórios...")
                atualizar_relatorio(nome_arquivo)

            # sair
            elif opcao == "6":
                print("\nEncerrando sistema...")
                print("Obrigado por usar o Sistema de Controle de Estoque!")
                break

            else:
                print("Opção inválida! Tente novamente.")

        # se der erro na digitacao
        except ValueError:
            print(f"ERRO: Valor inválido digitado. Por favor, tente novamente.")

# executar o programa
if __name__ == '__main__':
    menu_principal()
